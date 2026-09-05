from playwright.sync_api import sync_playwright
import re
import time
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook

URL = "https://www.tajhotels.com/en-in/bookings/landing-page?hotelId=d21c3bf6-f508-47ae-a456-540429b02b0d"

CHECK_EVERY_MINUTES = 30

ROOM_NAMES = [
    "SUPERIOR ROOM TWIN BED",
    "SUPERIOR ROOM KING BED"
]

TELEGRAM_CHAT_ID = "348797661"


def money_to_number(text):
    if not text:
        return None

    m = re.search(r"₹\s*([\d,]+)", text)

    if not m:
        return None

    return int(m.group(1).replace(",", ""))


def create_browser(p):

    print("Fresh browser open kar raha hoon...")

    browser = p.chromium.launch(
        headless=True,
        args=["--deny-permission-prompts"]
    )

    context = browser.new_context(
        viewport={
            "width": 1400,
            "height": 900
        },
        permissions=[]
    )

    context.add_init_script("""
        (() => {

            navigator.geolocation.getCurrentPosition =
                function(success, error) {
                    if (error) {
                        error({
                            code: 1,
                            message: "User denied Geolocation"
                        });
                    }
                };

            navigator.geolocation.watchPosition =
                function() {
                    return 0;
                };

        })();
    """)

    page = context.new_page()

    return browser, context, page


# ============================================================
# CALENDAR MONTH DETECTION
# ============================================================

def get_visible_months(page):

    result = []

    text = page.locator("body").inner_text()

    matches = re.findall(
        r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+2026",
        text,
        re.I
    )

    month_numbers = {
        "january": 1,
        "february": 2,
        "march": 3,
        "april": 4,
        "may": 5,
        "june": 6,
        "july": 7,
        "august": 8,
        "september": 9,
        "october": 10,
        "november": 11,
        "december": 12
    }

    for month in matches:

        value = (
            2026,
            month_numbers[month.lower()]
        )

        if value not in result:
            result.append(value)

    return result


# ============================================================
# OPEN CALENDAR
# ============================================================

def open_calendar(page):

    print("Date selector open kar raha hoon...")

    try:

        date_field = page.locator(
            "text=/\\d{1,2}\\s+[A-Za-z]{3}\\s+2026/"
        ).first

        date_field.click(
            timeout=5000
        )

    except:

        try:
            page.mouse.click(
                600,
                125
            )
        except:
            pass

    page.wait_for_timeout(
        1500
    )


# ============================================================
# NEXT MONTH
# ============================================================

def go_to_december(page):

    for attempt in range(12):

        months = get_visible_months(page)

        print(
            "Visible calendar:",
            months
        )

        if (
            2026,
            12
        ) in months:

            print(
                "December 2026 calendar mil gaya."
            )

            return True

        # ----------------------------------------------------
        # Find the actual right arrow from its position
        # ----------------------------------------------------

        clicked = False

        buttons = page.locator(
            "button, [role='button']"
        )

        for i in range(
            buttons.count()
        ):

            try:

                btn = buttons.nth(i)

                if not btn.is_visible():
                    continue

                box = btn.bounding_box()

                if not box:
                    continue

                aria = (
                    btn.get_attribute(
                        "aria-label"
                    ) or ""
                ).lower()

                title = (
                    btn.get_attribute(
                        "title"
                    ) or ""
                ).lower()

                combined = (
                    aria
                    + " "
                    + title
                )

                # Right side of calendar
                if (
                    box["x"] > 900
                    and box["y"] > 50
                    and box["y"] < 350
                    and (
                        "next" in combined
                        or "right" in combined
                        or "forward" in combined
                    )
                ):

                    btn.click(
                        timeout=3000
                    )

                    clicked = True

                    print(
                        "Next month click:",
                        attempt + 1
                    )

                    print(
                        "5 seconds wait..."
                    )

                    page.wait_for_timeout(
                        5000
                    )

                    break

            except:
                pass

        # ----------------------------------------------------
        # Coordinate fallback
        # ----------------------------------------------------

        if not clicked:

            try:

                # Browser viewport coordinate
                page.mouse.click(
                    1110,
                    115
                )

                print(
                    "Next arrow fallback click:",
                    attempt + 1
                )

                print(
                    "5 seconds wait..."
                )

                page.wait_for_timeout(
                    5000
                )

            except:
                pass

    return False


# ============================================================
# FIND DECEMBER CALENDAR AREA
# ============================================================

def get_december_area(page):

    # December heading locate karo
    headings = page.get_by_text(
        re.compile(
            r"^DECEMBER\s+2026$",
            re.I
        )
    )

    for i in range(
        headings.count()
    ):

        try:

            heading = headings.nth(i)

            if not heading.is_visible():
                continue

            box = heading.bounding_box()

            if not box:
                continue

            print(
                "December heading position:",
                round(box["x"], 1),
                round(box["y"], 1)
            )

            # ------------------------------------------------
            # Heading ke neeche calendar area.
            # Screenshot ke hisaab se:
            # December calendar right side par hota hai.
            # ------------------------------------------------

            return {
                "left": box["x"] - 120,
                "right": box["x"] + 500,
                "top": box["y"] + 30,
                "bottom": box["y"] + 420
            }

        except:
            pass

    return None


# ============================================================
# CLICK EXACT DECEMBER DATE
# ============================================================

def click_december_date(page, day):

    print(
        f"December 2026 ka {day} select kar raha hoon..."
    )

    area = get_december_area(page)

    if not area:

        print(
            "December calendar area nahi mila."
        )

        return False

    candidates = []

    # Exact text 25 / 26
    locator = page.get_by_text(
        str(day),
        exact=True
    )

    for i in range(
        locator.count()
    ):

        try:

            element = locator.nth(i)

            if not element.is_visible():
                continue

            box = element.bounding_box()

            if not box:
                continue

            x = box["x"] + box["width"] / 2
            y = box["y"] + box["height"] / 2

            # ------------------------------------------------
            # ONLY DECEMBER AREA
            # ------------------------------------------------

            if x < area["left"]:
                continue

            if x > area["right"]:
                continue

            if y < area["top"]:
                continue

            if y > area["bottom"]:
                continue

            candidates.append(
                (
                    x,
                    y,
                    element
                )
            )

        except:
            pass

    print(
        f"December ke andar {day} ke candidates:",
        len(candidates)
    )

    if not candidates:

        print(
            f"December 2026 ka {day} nahi mila."
        )

        return False

    # First candidate inside December area
    element = candidates[0][2]

    # --------------------------------------------------------
    # Real button / clickable parent
    # --------------------------------------------------------

    target = element

    try:

        parent = element.locator(
            "xpath=ancestor::button[1]"
        )

        if parent.count() > 0:

            if parent.first.is_visible():

                target = parent.first

    except:
        pass

    # --------------------------------------------------------
    # Click
    # --------------------------------------------------------

    try:

        target.scroll_into_view_if_needed()

    except:
        pass

    try:

        target.click(
            timeout=5000
        )

    except Exception as e:

        print(
            "Normal click failed:",
            e
        )

        try:

            box = target.bounding_box()

            if box:

                page.mouse.click(
                    box["x"] + box["width"] / 2,
                    box["y"] + box["height"] / 2
                )

            else:

                return False

        except Exception as e2:

            print(
                "Fallback click failed:",
                e2
            )

            return False

    page.wait_for_timeout(
        2000
    )

    print(
        f"{day} December 2026 click ho gaya."
    )

    return True


# ============================================================
# READ DATE FIELDS DIRECTLY
# ============================================================

def read_date_fields(page):

    values = []

    # Input fields
    inputs = page.locator(
        "input"
    )

    for i in range(
        inputs.count()
    ):

        try:

            inp = inputs.nth(i)

            if not inp.is_visible():
                continue

            value = (
                inp.get_attribute("value")
                or ""
            ).strip()

            placeholder = (
                inp.get_attribute("placeholder")
                or ""
            ).strip()

            if value:
                values.append(value)

            elif placeholder:
                values.append(placeholder)

        except:
            pass

    # Body text bhi check
    try:

        body = page.locator(
            "body"
        ).inner_text()

        matches = re.findall(
            r"\d{1,2}\s+[A-Za-z]{3}\s+2026",
            body
        )

        values.extend(matches)

    except:
        pass

    return values


# ============================================================
# SELECT DATES
# ============================================================

def select_dates(page):

    open_calendar(
        page
    )

    if not go_to_december(
        page
    ):

        print(
            "December 2026 calendar nahi mila."
        )

        return False

    # --------------------------------------------------------
    # 25 DECEMBER
    # --------------------------------------------------------

    if not click_december_date(
        page,
        25
    ):

        return False

    # --------------------------------------------------------
    # IMPORTANT
    # 25 click ke baad Taj calendar ko dobara open
    # kar sakte hain agar woh close ho gaya ho.
    # --------------------------------------------------------

    page.wait_for_timeout(
        2000
    )

    # Check if December calendar still visible
    months_after_25 = get_visible_months(
        page
    )

    print(
        "25 ke baad visible calendar:",
        months_after_25
    )

    if (
        2026,
        12
    ) not in months_after_25:

        print(
            "25 ke baad calendar close ho gaya."
        )

        print(
            "26 ke liye calendar dobara open kar raha hoon..."
        )

        open_calendar(
            page
        )

        if not go_to_december(
            page
        ):

            return False

    # --------------------------------------------------------
    # 26 DECEMBER
    # --------------------------------------------------------

    if not click_december_date(
        page,
        26
    ):

        return False

    page.wait_for_timeout(
        3000
    )

    # --------------------------------------------------------
    # DATE VALUES PRINT
    # --------------------------------------------------------

    values = read_date_fields(
        page
    )

    print(
        "Date fields:",
        values
    )

    # --------------------------------------------------------
    # IMPORTANT:
    # Agar header/input me December dates aa gayi hain
    # to confirmation.
    # --------------------------------------------------------

    combined = " ".join(
        values
    ).lower()

    if (
        "25 dec 2026" in combined
        and
        "26 dec 2026" in combined
    ):

        print(
            "25 Dec 2026 -> 26 Dec 2026 CONFIRMED."
        )

    else:

        # Taj ke UI me dates kabhi text ki jagah
        # selected state me ho sakti hain.
        # Isliye yahan SEARCH ko block nahi karenge.
        print(
            "Date fields me exact text nahi mila,"
            " lekin December ke exact 25/26 click kiye gaye hain."
        )

    return True


# ============================================================
# TELEGRAM
# ============================================================

def send_telegram(message):
    import os
    import urllib.parse
    import urllib.request

    token = os.environ.get("TAJ_TELEGRAM_BOT_TOKEN", "").strip()

    if not token:
        print("Telegram token nahi mila. TAJ_TELEGRAM_BOT_TOKEN set karo.")
        return False

    url = "https://api.telegram.org/bot" + token + "/sendMessage"
    data = urllib.parse.urlencode({
        "chat_id": TELEGRAM_CHAT_ID,
        "text": message
    }).encode("utf-8")

    try:
        req = urllib.request.Request(url, data=data, method="POST")
        with urllib.request.urlopen(req, timeout=20) as response:
            response.read()
        print("Telegram message sent.")
        return True
    except Exception as e:
        print("Telegram send error:", e)
        return False


# ============================================================
# EXCEL PRICE HISTORY
# ============================================================

def save_price_history(results, lowest, lowest_room, lowest_type):
    file_path = Path("Taj_Price_History.xlsx")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if file_path.exists():
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Price History"
        ws.append([
            "Date & Time",
            "Twin Member",
            "Twin Standard",
            "King Member",
            "King Standard",
            "Lowest Price",
            "Lowest Room",
            "Lowest Rate Type"
        ])

    twin = results.get("SUPERIOR ROOM TWIN BED") or {}
    king = results.get("SUPERIOR ROOM KING BED") or {}

    ws.append([
        now,
        twin.get("member"),
        twin.get("standard"),
        king.get("member"),
        king.get("standard"),
        lowest,
        lowest_room,
        lowest_type.upper()
    ])

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 18

    wb.save(file_path)
    print("Excel history saved:", file_path)


# ============================================================
# ROOM HEADING
# ============================================================

def find_room_heading(page, room_name):
    loc = page.get_by_text(room_name, exact=True)

    for i in range(loc.count()):
        try:
            el = loc.nth(i)
            if el.is_visible():
                return el
        except Exception:
            pass

    # Fallback: text may have line breaks / extra spaces.
    loc = page.get_by_text(room_name, exact=False)
    for i in range(min(loc.count(), 20)):
        try:
            el = loc.nth(i)
            if el.is_visible() and room_name.upper() in el.inner_text().upper():
                return el
        except Exception:
            pass

    return None


def is_100_percent_nonrefundable(text):
    upper = re.sub(r"\s+", " ", text.upper())
    has_100 = (
        "100 PCT" in upper
        or "100%" in upper
        or "100 PERCENT" in upper
    )
    has_nonref = (
        "NON-REFUNDABLE" in upper
        or "NON REFUNDABLE" in upper
        or "NONREFUNDABLE" in upper
    )
    return has_100 and has_nonref


def extract_prices(text):
    values = []
    for item in re.findall(r"₹\s*[\d,]+", text):
        value = money_to_number(item)
        if value is not None and value not in values:
            values.append(value)
    return values


def _room_section_text(page, room_name):
    """Get rendered text belonging to the requested room."""
    body = page.locator("body").inner_text(timeout=10000)
    text = re.sub(r"\s+", " ", body).strip()
    upper = text.upper()
    target = room_name.upper()
    other = ("SUPERIOR ROOM KING BED" if "TWIN" in target
             else "SUPERIOR ROOM TWIN BED")

    starts = [m.start() for m in re.finditer(re.escape(target), upper)]
    best = None

    for pos in starts:
        tail = text[pos:]
        other_pos = tail.upper().find(other)
        block = tail if other_pos < 0 else tail[:other_pos]
        price_count = len(re.findall(r"₹\s*[\d,]+", block))
        rate_count = len(re.findall(r"\b(?:MEMBER RATE|STANDARD RATE)\b", block, re.I))
        if price_count or rate_count:
            score = (price_count, rate_count, -len(block))
            if best is None or score > best[0]:
                best = (score, block)

    return best[1] if best else None


def _rate_events(section):
    """Read each MEMBER/STANDARD rate and its immediately preceding card text."""
    pattern = re.compile(r"\b(MEMBER RATE|STANDARD RATE)\b", re.I)
    matches = list(pattern.finditer(section))
    events = []

    for idx, match in enumerate(matches):
        label = match.group(1).upper()
        previous_label_pos = matches[idx - 1].start() if idx > 0 else 0
        segment = section[previous_label_pos:match.end() + 350]

        price_match = re.search(
            r"₹\s*([\d,]+)",
            section[match.end():match.end() + 350]
        )
        if not price_match:
            continue

        events.append({
            "label": "member" if "MEMBER" in label else "standard",
            "price": int(price_match.group(1).replace(",", "")),
            "segment": segment,
        })

    return events


def extract_room_rates(page, room_name):
    print("")
    print(room_name + ": actual displayed room prices read kar raha hoon...")

    try:
        section = _room_section_text(page, room_name)
        if not section:
            print(room_name + ": room ka rendered text nahi mila.")
            return None

        events = _rate_events(section)
        if not events:
            print(room_name + ": actual ₹ price nahi mila.")
            return None

        member = None
        standard = None
        skip_standard_after_rejected_member = False

        for event in events:
            label = event["label"]
            price = event["price"]

            # A 100% non-refundable condition belongs to the rate card whose
            # first rate follows it. If that card also exposes a STANDARD RATE,
            # skip that standard rate until the next MEMBER RATE starts.
            if label == "member":
                skip_standard_after_rejected_member = False

                if is_100_percent_nonrefundable(event["segment"]):
                    print("MEMBER RATE: 100% full-stay non-refundable card ignore kiya.")
                    skip_standard_after_rejected_member = True
                    continue

                if member is None:
                    member = price
                    print(room_name + " MEMBER RATE: ₹{:,.0f}".format(member))

            else:
                if skip_standard_after_rejected_member:
                    print("STANDARD RATE: same 100% non-refundable card ignore kiya.")
                    continue

                if is_100_percent_nonrefundable(event["segment"]):
                    print("STANDARD RATE: 100% full-stay non-refundable card ignore kiya.")
                    continue

                if standard is None:
                    standard = price
                    print(room_name + " STANDARD RATE: ₹{:,.0f}".format(standard))

            if member is not None and standard is not None:
                break

        if member is None and standard is None:
            print(room_name + ": actual price nahi mila.")
            return None

        return {"member": member, "standard": standard}

    except Exception as e:
        print(room_name + ": rate extraction error:", e)
        return None


def scroll_to_king(page):
    print("SUPERIOR ROOM KING BED ko actual mouse-wheel se screen par la raha hoon...")

    page.mouse.move(700, 650)

    for step in range(1, 16):
        heading = find_room_heading(page, "SUPERIOR ROOM KING BED")
        if heading:
            box = heading.bounding_box()
            if box:
                vh = page.evaluate("window.innerHeight")
                if 120 <= box["y"] <= vh - 120:
                    print("KING BED screen par aa gaya.")
                    return True

        before = page.evaluate("window.scrollY")
        page.mouse.wheel(0, 700)
        page.wait_for_timeout(700)
        after = page.evaluate("window.scrollY")
        print("Mouse-wheel step:", step, "windowY:", round(before, 1), "->", round(after, 1))

        # Taj's room list can be inside an internal scrollable element.
        if after == before:
            moved = page.evaluate("""() => {
                const nodes = Array.from(document.querySelectorAll('*'));
                const candidates = nodes.filter(e => {
                    const s = getComputedStyle(e);
                    return (s.overflowY === 'auto' || s.overflowY === 'scroll') &&
                           e.scrollHeight > e.clientHeight + 30;
                });
                let best = null;
                let bestTop = Infinity;
                for (const e of candidates) {
                    const r = e.getBoundingClientRect();
                    if (r.width < 300 || r.height < 200) continue;
                    const top = Math.abs(r.top - 100);
                    if (top < bestTop) { bestTop = top; best = e; }
                }
                if (!best) return {changed:false};
                const beforeTop = best.scrollTop;
                best.scrollTop = Math.min(beforeTop + 700, best.scrollHeight - best.clientHeight);
                return {changed: best.scrollTop !== beforeTop, before: beforeTop, after: best.scrollTop};
            }""")
            print("Internal scroll:", moved)

        page.wait_for_timeout(500)

    print("King heading ko limited scrolling ke baad nahi la paya.")
    return False

# ============================================================
# ONE CHECK
# ============================================================


# ============================================================
# TAJ NETWORK/API PRICE CHECK
# ============================================================

TAJ_API_HOST = "api-cug1-825v2.tajhotels.com"
TAJ_API_PATH = "/hudiniService/v1/hotel-availability"

TARGET_API_PAYLOAD = {
    "endDate": "2026-12-26",
    "numRooms": 1,
    "adults": 1,
    "children": 0,
    "startDate": "2026-12-25",
    "hotelId": "d21c3bf6-f508-47ae-a456-540429b02b0d",
    "rateFilter": "RRM,PKG,MD",
    "memberTier": "member",
    "package": "PKG",
    "isOfferLandingPage": False,
    "rateCode": None,
    "promoCode": None,
    "promoType": None,
    "couponCode": None,
    "agentId": None,
    "agentType": None,
    "isMyAccount": False,
    "isCorporate": False,
    "isLogin": False,
    "isMemberOffer1": False,
    "isMemberOffer2": False,
    "forSomeoneElse": False,
    "isEmployeeOffer": False
}

TARGET_ROOM_CODES = {
    "DTX": "SUPERIOR ROOM TWIN BED",
    "DKX": "SUPERIOR ROOM KING BED"
}

def api_price(rate):
    try:
        return float(rate["daily"][0]["price"]["amount"])
    except Exception:
        return None

def rate_is_100pct_nonrefundable(rate):
    parts = []
    rc = rate.get("rateContent") or {}
    details = rc.get("details") or {}
    bp = rate.get("bookingPolicy") or {}
    parts += [
        rc.get("name"), details.get("description"),
        details.get("detailedDescription"), details.get("displayName"),
        details.get("displayDescription"), bp.get("description"),
        bp.get("refundableStay")
    ]
    text = re.sub(r"\s+", " ", " ".join(str(x) for x in parts if x)).upper()
    return (
        ("100 PCT" in text or "100%" in text or "100 PERCENT" in text)
        and ("NON-REFUNDABLE" in text or "NON REFUNDABLE" in text or "NONREFUNDABLE" in text)
    )

def get_taj_api_response(page, context):
    captured = {"headers": None, "url": None, "body": None}

    def capture_request(req):
        if TAJ_API_HOST in req.url:
            if captured["headers"] is None:
                captured["headers"] = req.all_headers()
                captured["url"] = req.url
                captured["body"] = req.post_data

    page.on("request", capture_request)

    # Loading the normal Taj page makes Taj itself issue a legitimate API
    # request, establishing the browser cookies/session and giving us the exact
    # browser headers required by the security layer.
    page.goto(URL, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(12000)

    if not captured["headers"]:
        raise RuntimeError("Taj API request browser se capture nahi hui.")

    print("Taj browser API request captured.")
    print("Original API URL:", captured["url"])

    headers = dict(captured["headers"])
    for key in list(headers):
        low = key.lower()
        # Browser request headers can contain HTTP/2 pseudo-headers such as
        # :authority. Playwright APIRequestContext rejects those, so remove
        # pseudo-headers and browser-managed headers.
        if low.startswith(":") or low in (
            "content-length", "host", "cookie", "origin", "referer"
        ):
            headers.pop(key, None)

    headers["content-type"] = "application/json"
    headers["accept"] = "application/json"

    api_url = "https://" + TAJ_API_HOST + TAJ_API_PATH

    # context.request shares the browser context's cookies with Playwright's
    # API client, while using the same browser-derived headers.
    response = context.request.post(
        api_url,
        data=json.dumps(TARGET_API_PAYLOAD),
        headers=headers,
        timeout=60000
    )

    print("TAJ API HTTP:", response.status)

    body = response.text()
    if response.status != 200:
        raise RuntimeError(
            "Taj API returned HTTP " + str(response.status) +
            ": " + body[:1000]
        )

    data = json.loads(body)

    Path("taj_last_api_response.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    return data

def extract_api_rates(data):
    results = {
        "SUPERIOR ROOM TWIN BED": {},
        "SUPERIOR ROOM KING BED": {}
    }

    room_rates = data.get("roomAvailability", {}).get("roomRates", [])

    for room in room_rates:
        room_name = TARGET_ROOM_CODES.get(room.get("roomCode"))
        if not room_name:
            continue

        member = None
        standard = None

        for rate in room.get("rooms", []):
            if rate_is_100pct_nonrefundable(rate):
                print(
                    room_name, "|", rate.get("rateCode"),
                    "-> 100% full-stay non-refundable EXCLUDED"
                )
                continue

            if member is None:
                member = api_price(rate.get("memberRate"))
            if standard is None:
                standard = api_price(rate.get("standardRate"))

            if member is not None and standard is not None:
                break

        results[room_name] = {
            "member": member,
            "standard": standard
        }

        print(
            room_name,
            "| Member:", member,
            "| Standard:", standard
        )

    return results


def check_price(p):
    browser = context = page = None

    try:
        print("\n" + "=" * 65)
        print("CHECK TIME:", datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
        print("Hotel     : Taj City Centre Gurugram")
        print("Dates     : 25 Dec 2026 -> 26 Dec 2026")
        print("Guests    : 1 | Rooms: 1")
        print("Method    : TAJ BROWSER NETWORK API")
        print("=" * 65)

        browser, context, page = create_browser(p)

        results = extract_api_rates(
            get_taj_api_response(page, context)
        )

        valid_prices = []
        for room_name, data in results.items():
            for rate_type in ("member", "standard"):
                value = data.get(rate_type)
                if value is not None:
                    valid_prices.append((value, room_name, rate_type))

        if not valid_prices:
            print("Koi valid price detect nahi hua.")
            return None

        valid_prices.sort(key=lambda x: x[0])
        lowest, lowest_room, lowest_type = valid_prices[0]

        lines = [
            "💰 ₹{:,.0f} / NIGHT".format(lowest),
            "LOWEST PRICE", "",
            "🏨 Taj City Centre Gurugram",
            "📅 25 Dec 2026 → 26 Dec 2026",
            "👤 1 Guest | 1 Room", ""
        ]

        for room_name in ROOM_NAMES:
            d = results.get(room_name) or {}
            lines.append("🛏 " + room_name)
            lines.append(
                "Member Rate: ₹{:,.0f} / night".format(d["member"])
                if d.get("member") is not None
                else "Member Rate: Not available"
            )
            lines.append(
                "Standard Rate: ₹{:,.0f} / night".format(d["standard"])
                if d.get("standard") is not None
                else "Standard Rate: Not available"
            )
            lines.append("")

        lines += [
            "❌ 100% full-stay non-refundable rate cards excluded",
            "🔗 " + URL
        ]

        telegram_message = "\n".join(lines)

        print("\n" + "=" * 65)
        print("LOWEST PRICE : ₹{:,.0f} / night".format(lowest))
        print("ROOM         :", lowest_room)
        print("RATE TYPE    :", lowest_type.upper())
        print("=" * 65)
        print(telegram_message)

        save_price_history(results, lowest, lowest_room, lowest_type)
        send_telegram(telegram_message)

        return lowest

    except Exception as e:
        print("\nERROR:", e)
        print("Current API check fail hua.")
        return None

    finally:
        for obj in (page, context, browser):
            try:
                if obj:
                    obj.close()
            except Exception:
                pass
        print("Current check ka browser close kar diya gaya.")


# ============================================================
# SINGLE CLOUD CHECK
# ============================================================

with sync_playwright() as p:
    print("\n" + "=" * 65)
    print("TAJ CITY CENTRE GURUGRAM PRICE TRACKER - CLOUD CHECK")
    print("=" * 65)
    print("Dates        : 25 Dec 2026 -> 26 Dec 2026")
    print("Rooms        : Twin Bed + King Bed")
    print("This run    : One check")
    print("Telegram     : Enabled")
    print("=" * 65)

    try:
        check_price(p)
    except Exception as e:
        print("\nCLOUD CHECK ERROR:", e)
        raise
