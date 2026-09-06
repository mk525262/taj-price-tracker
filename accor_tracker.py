from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
from datetime import datetime
from pathlib import Path
import os
import requests

HOTEL_ID = "6529"
HOTEL_NAME = "ibis Jaipur City Centre"
CHECKIN = "2026-12-25"
CHECKOUT = "2026-12-26"
ADULTS = 1
ROOMS = 1
HOTEL_URL = f"https://all.accor.com/hotel/{HOTEL_ID}/index.en.shtml"
HISTORY_FILE = Path("Accor_Ibis_Jaipur_Price_History.xlsx")
TELEGRAM_CHAT_ID = "348797661"


def send_telegram(message):
    token = os.environ.get("TAJ_TELEGRAM_BOT_TOKEN")
    if not token:
        raise RuntimeError("TAJ_TELEGRAM_BOT_TOKEN secret is missing")
    r = requests.post(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data={"chat_id": TELEGRAM_CHAT_ID, "text": message},
        timeout=30,
    )
    r.raise_for_status()


def save_history(rows):
    if HISTORY_FILE.exists():
        wb = load_workbook(HISTORY_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Price History"
        ws.append([
            "Check Time", "Check-in", "Check-out", "Hotel", "Room",
            "Member Price (INR)", "Standard Price (INR)", "Rate", "Meal Plan",
            "Cancellation", "Guarantee", "Eligible"
        ])
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in rows:
        ws.append([
            now, CHECKIN, CHECKOUT, HOTEL_NAME, row["room"],
            row["member_price_inr"], row["standard_price_inr"], row["rate"], row["meal"],
            row["cancellation"], row["guarantee"], "YES" if row["eligible"] else "NO"
        ])
    wb.save(HISTORY_FILE)
    print(f"Excel history saved: {HISTORY_FILE}", flush=True)


def find_value(obj, wanted_key):
    if isinstance(obj, dict):
        if wanted_key in obj and isinstance(obj[wanted_key], (int, float)):
            return obj[wanted_key]
        for value in obj.values():
            found = find_value(value, wanted_key)
            if found is not None:
                return found
    elif isinstance(obj, list):
        for value in obj:
            found = find_value(value, wanted_key)
            if found is not None:
                return found
    return None


def main():
    hot = None
    cold = None
    basket_inr = None

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(locale="en-IN")

        def on_response(resp):
            nonlocal hot, cold, basket_inr
            try:
                url = resp.url
                if "api.accor.com/bff/v1/graphql" in url:
                    data = resp.json()
                    if not isinstance(data, dict):
                        return
                    hotel_offers = data.get("data", {}).get("hotelOffers", {})
                    offers = hotel_offers.get("offersSelection", {}).get("offers")
                    if offers:
                        hot = data
                        print(f"HotelPageHot captured: {len(offers)} offers", flush=True)
                    accommodations = data.get("data", {}).get("hotel", {}).get("accommodations")
                    if accommodations is not None:
                        cold = data
                # Basket may be served by a separate Accor endpoint.
                if "basket" in url.lower():
                    try:
                        data = resp.json()
                        basket_total = find_value(data, "totalAmountHotelCurrency")
                        if basket_total is not None:
                            basket_inr = basket_total
                            print(f"Accor basket INR total captured: ₹{basket_inr:,.2f}", flush=True)
                    except Exception:
                        pass
            except Exception:
                pass

        context.on("response", on_response)
        page = context.new_page()
        target_url = HOTEL_URL + f"?dateIn={CHECKIN}&dateOut={CHECKOUT}&compositions=1&stayplus=false"
        print("Method : ACCOR BROWSER GRAPHQL API", flush=True)
        print("Opening Accor booking page...", flush=True)
        page.goto(target_url, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(12000)

        if not hot:
            try:
                loc = page.get_by_text("See availabilities", exact=False).first
                if loc.is_visible(timeout=1500):
                    print("Triggering: See availabilities", flush=True)
                    loc.evaluate("el => el.click()")
                    page.wait_for_timeout(20000)
            except Exception as e:
                print("Trigger click failed:", str(e)[:300], flush=True)

        if not hot:
            raise RuntimeError("Accor HotelPageHot response was not captured")

        offers = hot["data"]["hotelOffers"]["offersSelection"]["offers"]
        room_names = {}
        if cold:
            for acc in cold["data"]["hotel"]["accommodations"]:
                if acc.get("code"):
                    room_names[acc["code"]] = acc.get("name") or acc["code"]

        parsed = []
        for offer in offers:
            if offer.get("type") != "ROOM":
                continue
            pricing = offer.get("pricing", {}) or {}
            main = pricing.get("main", {}) or {}
            alternative = pricing.get("alternative", {}) or {}
            rate = offer.get("rate", {}) or {}
            meal = offer.get("mealPlan", {}) or {}
            policies = main.get("simplifiedPolicies", {}) or {}
            cancellation = (policies.get("cancellation") or {}).get("label") or ""
            guarantee = (policies.get("guarantee") or {}).get("label") or ""
            member = "MEMBER_RATE" in (main.get("categories") or [])
            eligible = not (
                cancellation.lower() == "non-refundable"
                and guarantee.lower() in {"online payment", "prepaid"}
            )
            product_id = (offer.get("product") or {}).get("id")
            parsed.append({
                "product": product_id,
                "room": room_names.get(product_id, product_id or "ROOM"),
                "member": member,
                "member_price_eur": main.get("amount"),
                "standard_price_eur": alternative.get("amount") if alternative.get("categories") == ["STANDARD"] else None,
                "currency": pricing.get("currency"),
                "rate": rate.get("label") or rate.get("id"),
                "meal": meal.get("label") or meal.get("code") or "Room only",
                "cancellation": cancellation,
                "guarantee": guarantee,
                "eligible": eligible,
            })

        eligible_members = [x for x in parsed if x["member"] and x["eligible"] and isinstance(x["member_price_eur"], (int, float))]
        if not eligible_members:
            raise RuntimeError("No eligible Accor member room offer found")

        first_offer_eur = next((x["member_price_eur"] for x in parsed if isinstance(x["member_price_eur"], (int, float))), None)
        if not basket_inr or not first_offer_eur:
            raise RuntimeError("Could not capture Accor basket INR conversion")
        eur_to_inr = basket_inr / first_offer_eur
        print(f"Live Accor EUR→INR factor: {eur_to_inr:.6f}", flush=True)

        best_by_room = {}
        for row in eligible_members:
            row["member_price_inr"] = round(row["member_price_eur"] * eur_to_inr, 2)
            row["standard_price_inr"] = round(row["standard_price_eur"] * eur_to_inr, 2) if isinstance(row["standard_price_eur"], (int, float)) else None
            key = row["product"]
            if key not in best_by_room or row["member_price_inr"] < best_by_room[key]["member_price_inr"]:
                best_by_room[key] = row

        final_rows = sorted(best_by_room.values(), key=lambda x: x["member_price_inr"])
        lowest = final_rows[0]
        print("\nACCOR ELIGIBLE ROOMS:", flush=True)
        for row in final_rows:
            standard = f"₹{row['standard_price_inr']:,.0f}" if row["standard_price_inr"] is not None else "N/A"
            print(f"{row['room']} | Member ₹{row['member_price_inr']:,.0f} | Standard {standard} | {row['rate']} | {row['meal']}", flush=True)
        print(f"LOWEST PRICE : ₹{lowest['member_price_inr']:,.0f} / night", flush=True)
        print(f"ROOM : {lowest['room']}", flush=True)
        print("RATE TYPE : MEMBER", flush=True)

        save_history(final_rows)
        msg = f"🏨 ACCOR PRICE UPDATE\n\n{HOTEL_NAME}\n📅 {CHECKIN} → {CHECKOUT}\n👤 {ADULTS} Adult | {ROOMS} Room\n\n"
        for row in final_rows:
            standard = f"₹{row['standard_price_inr']:,.0f}" if row["standard_price_inr"] is not None else "N/A"
            msg += f"{row['room']}\nMember: ₹{row['member_price_inr']:,.0f}\nStandard: {standard}\n\n"
        msg += f"Lowest eligible: ₹{lowest['member_price_inr']:,.0f} ({lowest['room']})"
        send_telegram(msg)
        print("Telegram message sent.", flush=True)
        browser.close()


if __name__ == "__main__":
    main()
